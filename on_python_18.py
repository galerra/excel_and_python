import openpyxl
wookbook = openpyxl.load_workbook("9_18.xlsx")
worksheet = wookbook.active
main_list = []

for i in range(1, worksheet.max_row):
    count = 0
    for col in worksheet.iter_cols(2, worksheet.max_column):
        if (isinstance(col[i].value, int)) or (isinstance(col[i].value, float)):
            count += col[i].value
    if count != 0:
        count = count / 24
        count = round(count,3)
        main_list.append(count)
print(main_list)

count = 0
for i in range(1, len(main_list)):
    for col in worksheet.iter_cols(10, 10):
        if col[i].value > main_list[i - 1]:
            count += 1
print(count)